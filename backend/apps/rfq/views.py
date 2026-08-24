"""
RFQ ViewSets — CRUD for rfq_master, customers, estimators, goals, and print view.
"""
import re
import math
import io
import datetime as dt
from datetime import datetime, date, time
from decimal import Decimal, InvalidOperation
from rest_framework import viewsets, generics, filters, status

def coerce_bool(val):
    if val is None: return False
    if isinstance(val, bool): return val
    return str(val).strip().upper() in ('Y', 'YES', 'TRUE', '1', 'X')

def coerce_date(val):
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        val = val.strip()
        if not val: return None
        # Add support for word-based month formats (e.g. 5 July 2026)
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%d-%b-%Y', '%d %B %Y', '%d %b %Y', '%B %d, %Y', '%b %d, %Y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    return None

def coerce_decimal(val):
    if val is None: return None
    if isinstance(val, dt.timedelta):
        return Decimal(str(round(val.total_seconds() / 3600.0, 2)))
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (val != val or math.isnan(val)):
            return None
        return Decimal(str(val))
    try:
        return Decimal(str(val).replace(',', '').strip())
    except (InvalidOperation, ValueError, TypeError):
        return None

def coerce_int(val):
    if val is None: return None
    try:
        f_val = float(str(val).replace(',', ''))
        if math.isnan(f_val):
            return None
        return int(f_val)
    except (ValueError, TypeError):
        return None

def coerce_time(val):
    if val is None: return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        if not val: return None
        for fmt in ('%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M%p'):
            try:
                return datetime.strptime(val, fmt).time()
            except ValueError:
                pass
    return None
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as df_filters
from django.db import models, transaction
from django.db.models import Q
from django.core.mail import send_mail
from django.conf import settings
import os
import html
import threading
import logging
import quote_helper

from .models import RFQMaster, Estimator, MonthlyBidGoal, SystemSetting, QuoteWorkflow
from projects.models import Customer, CustomerContact
from .serializers import (
    RFQMasterSerializer, RFQListSerializer, PrintSetupSerializer,
    CustomerSerializer, EstimatorSerializer, MonthlyBidGoalSerializer,
    SystemSettingSerializer, QuoteWorkflowSerializer,
)
from .permissions import CanEditRFQ, IsManagerOrReadOnly
logger = logging.getLogger(__name__)


def decode_mime_header(header_val):
    """Safely decodes RFC 2047 MIME-encoded headers combining all fragments."""
    if not header_val:
        return ""
    try:
        from email.header import decode_header
        parts = decode_header(header_val)
        decoded_str = ""
        for content, encoding in parts:
            if isinstance(content, bytes):
                decoded_str += content.decode(encoding or "utf-8", errors="replace")
            else:
                decoded_str += str(content)
        return decoded_str.strip()
    except Exception:
        return str(header_val).strip()


def extract_clean_email(raw_addr):
    """Extracts raw email address from complex header string (e.g. 'Name <email@dom.com>')."""
    if not raw_addr:
        return ""
    m = re.search(r'<([^>]+)>', raw_addr)
    if m:
        return m.group(1).strip()
    m = re.search(r'[\w\.\+\-]+@[\w\.\-]+\.[a-zA-Z]{2,}', raw_addr)
    if m:
        return m.group(0).strip()
    return str(raw_addr).strip()



# ─────────────────────────────────────────────────────────────────────────────
# FILTERS
# ─────────────────────────────────────────────────────────────────────────────

class RFQFilter(df_filters.FilterSet):
    year = df_filters.NumberFilter(field_name='quote_date', lookup_expr='year')
    bid_year = df_filters.NumberFilter(field_name='bid_due_date', lookup_expr='year')
    customer = df_filters.NumberFilter(field_name='customer__id')
    estimator = df_filters.NumberFilter(field_name='primary_estimator__id')
    won_lost = df_filters.CharFilter(field_name='won_lost')
    bid_due_from = df_filters.DateFilter(field_name='bid_due_date', lookup_expr='gte')
    bid_due_to = df_filters.DateFilter(field_name='bid_due_date', lookup_expr='lte')
    is_active = df_filters.BooleanFilter(field_name='deleted_at', lookup_expr='isnull')
    budget_type = df_filters.CharFilter(field_name='budget_type')

    class Meta:
        model = RFQMaster
        fields = ['won_lost', 'budget_type', 'decision_to_bid']


# ─────────────────────────────────────────────────────────────────────────────
# RFQ MASTER VIEWSET
# ─────────────────────────────────────────────────────────────────────────────

def send_rfq_email_async(rfq_id, project_link=None):
    """Function to run in a background thread to send the email."""
    try:
        rfq = RFQMaster.objects.select_related('customer', 'primary_estimator').get(id=rfq_id)
        if not rfq.scope_of_work:
            return
            
        recipient_names = []
        recipient_emails = []
        
        detailing_emails = 'namrutha@caldimengg.in'
        fabrication_emails = 'divya@caldimengg.in'
        erection_emails = 'divya@caldimengg.in'
        try:
            detailing_emails = SystemSetting.objects.get(key='rfq_detailing_emails').value
        except SystemSetting.DoesNotExist:
            pass
        try:
            fabrication_emails = SystemSetting.objects.get(key='rfq_fabrication_emails').value
        except SystemSetting.DoesNotExist:
            pass
        try:
            erection_emails = SystemSetting.objects.get(key='rfq_erection_emails').value
        except SystemSetting.DoesNotExist:
            pass

        scopes = [s.strip() for s in rfq.scope_of_work.split(',') if s.strip()]
        
        def add_recipients(email_str):
            for e in email_str.split(','):
                e = e.strip()
                if e:
                    recipient_emails.append(e)
                    recipient_names.append(e.split('@')[0])

        if 'Detailing' in scopes and detailing_emails:
            add_recipients(detailing_emails)
        if 'Fabrication' in scopes and fabrication_emails:
            add_recipients(fabrication_emails)
        if 'Erection' in scopes and erection_emails:
            add_recipients(erection_emails)
            
        if not recipient_emails:
            return
            
        recipient_name = " and ".join(list(set(recipient_names)))
        recipient_emails = list(set(recipient_emails))
            
        import datetime
        
        subject = f"Project Details: {rfq.quote_no} - {rfq.project_name} [Quote Ref: {rfq.quote_no}]"
        
        # Determine estimator details
        estimator_name = "Estimator"
        estimator_phone = "717-464-0330"
        estimator_email = "estimator@steelfabenterprises.com"
        
        if rfq.primary_estimator:
            initials = (rfq.primary_estimator.initials or "").upper().strip()
            full_name = (rfq.primary_estimator.full_name or "").strip()
            
            contact_map = {
                'AS': {
                    'name': 'Andy Smith',
                    'phone': '717-464-0330 x223',
                    'email': 'asmith@steelfabenterprises.com'
                },
                'CR': {
                    'name': 'Chris R.',
                    'phone': '717-464-0330',
                    'email': 'estimator@steelfabenterprises.com'
                },
            }
            
            primary_initials = initials.split('/')[0].strip() if '/' in initials else initials
            if primary_initials in contact_map:
                c = contact_map[primary_initials]
                estimator_name = c['name']
                estimator_phone = c['phone']
                estimator_email = c['email']
            elif full_name:
                estimator_name = full_name
                parts = full_name.lower().split()
                if len(parts) >= 2:
                    estimator_email = f"{parts[0][0]}{parts[1]}@steelfabenterprises.com"
                else:
                    estimator_email = f"{parts[0]}@steelfabenterprises.com"

        # Format dates
        due_date_str = f"{rfq.bid_due_date.strftime('%B')} {rfq.bid_due_date.day}, {rfq.bid_due_date.year}" if rfq.bid_due_date else "N/A"
        
        deadline_date = rfq.bid_due_date - datetime.timedelta(days=3) if rfq.bid_due_date else None
        if deadline_date:
            if deadline_date.weekday() == 5:    # Saturday
                deadline_date -= datetime.timedelta(days=1)
            elif deadline_date.weekday() == 6:  # Sunday
                deadline_date -= datetime.timedelta(days=2)
        deadline_str = f"{deadline_date.strftime('%B')} {deadline_date.day}, {deadline_date.year}" if deadline_date else "N/A"

        from django.utils.html import escape
        
        esc_project_name = escape(rfq.project_name)
        esc_location = escape(rfq.location or 'N/A')
        esc_scope = escape(rfq.scope_of_work or 'project')
        esc_comments = escape(rfq.project_comments.strip()) if rfq.project_comments else ''
        
        # 1. Plain text body
        body = f"Good Morning –\n\n"
        body += f"Please see the link below for the {rfq.project_name} project located in {rfq.location or 'N/A'} "
        body += f"for your review in providing an updated Model and IFC files for the {rfq.scope_of_work or 'project'}. "
        
        if rfq.project_comments:
            body += f"{rfq.project_comments.strip()} "
            
        if rfq.budget_type == 'Rebid':
            prev_bid_month_year = rfq.quote_date.strftime('%B %Y') if rfq.quote_date else "previously"
            body += f"This project previously bid in {prev_bid_month_year} and is out for best and final offer bidding on {due_date_str}, "
        elif rfq.budget_type == 'Budget':
            body += f"This project is out for budget pricing on {due_date_str}, "
        else:
            body += f"This project is out for bidding on {due_date_str}, "
            
        body += f"if you could review and forward your updated files to estimator@steelfabenterprises.com by {deadline_str} would be greatly appreciated.\n\n"
        
        link_url = project_link if project_link else f"https://caldimproducts.com/rfq/data-entry?quote_no={rfq.quote_no}"
        body += f"{link_url}\n\n"
        body += "As a reminder, we will need Detailing Quote for the Structural & Miscellaneous and Engineering as well.\n\n"
        body += f"Should you have any questions or need any additional information, please contact {estimator_name} at {estimator_phone} or {estimator_email}. Thanks for your help!\n"

        # 2. HTML body
        html_body = f"Good Morning –<br/><br/>"
        html_body += f"Please see the link below for <b>the {esc_project_name} project</b> located in <b>{esc_location}</b> "
        html_body += f"for your review in providing an updated Model and IFC files for the <b><u>{esc_scope}</u></b>. "
        
        if esc_comments:
            html_body += f"{esc_comments} "
            
        if rfq.budget_type == 'Rebid':
            prev_bid_month_year = rfq.quote_date.strftime('%B %Y') if rfq.quote_date else "previously"
            html_body += f"This project previously bid in {prev_bid_month_year} and is out for best and final offer bidding on <b>{due_date_str}</b>, "
        elif rfq.budget_type == 'Budget':
            html_body += f"This project is out for budget pricing on <b>{due_date_str}</b>, "
        else:
            html_body += f"This project is out for bidding on <b>{due_date_str}</b>, "
            
        html_body += f"if you could review and forward your updated files to <a href=\"mailto:estimator@steelfabenterprises.com\">estimator@steelfabenterprises.com</a> by <b>{deadline_str}</b> would be greatly appreciated.<br/><br/>"
        
        html_body += f'<a href="{link_url}">{link_url}</a><br/><br/>'
        html_body += "<b><i>As a reminder, we will need Detailing Quote for the Structural & Miscellaneous and Engineering as well.</i></b><br/><br/>"
        
        # Link email contacts inside HTML
        esc_estimator_name = escape(estimator_name)
        esc_estimator_phone = escape(estimator_phone)
        esc_estimator_email = escape(estimator_email)
        html_body += f"Should you have any questions or need any additional information, please contact {esc_estimator_name} at {esc_estimator_phone} or <b><u><a href=\"mailto:{esc_estimator_email}\">{esc_estimator_email}</a></u></b>. Thanks for your help!<br/>"

        # Send separate individual emails to each recipient so vendors never see each other or get put in CC
        for r_email in recipient_emails:
            try:
                send_mail(
                    subject=subject,
                    message=body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[r_email],
                    fail_silently=False,
                    html_message=html_body
                )
            except Exception as e:
                logger.error(f"Failed to send RFQ notification to {r_email}: {e}")
        
        rfq.email_sent = True
        rfq.save(update_fields=['email_sent'])

        # Ensure QuoteWorkflow is created/updated
        customer_email = ""
        if rfq.customer:
            contact = rfq.customer.contacts.filter(email__isnull=False).exclude(email='').first()
            if contact and contact.email:
                customer_email = contact.email

        fab_erect_list = []
        if 'Fabrication' in scopes and fabrication_emails:
            fab_erect_list.extend([e.strip() for e in fabrication_emails.split(',') if e.strip()])
        if 'Erection' in scopes and erection_emails:
            fab_erect_list.extend([e.strip() for e in erection_emails.split(',') if e.strip()])
        
        est_fab_erect_str = ", ".join(list(set(fab_erect_list))) if fab_erect_list else estimator_email
        det_str = detailing_emails if 'Detailing' in scopes else ''

        workflow, created = QuoteWorkflow.objects.get_or_create(
            quote_id=rfq.quote_no,
            defaults={
                'rfq': rfq,
                'subject': subject,
                'sender': customer_email,
                'status': QuoteWorkflow.WorkflowStatus.FORWARDED,
                'estimator_email': est_fab_erect_str,
                'detailer_email': det_str,
            }
        )
        workflow.rfq = rfq
        workflow.subject = subject
        if customer_email:
            workflow.sender = customer_email
        workflow.estimator_email = est_fab_erect_str
        workflow.detailer_email = det_str
        # If workflow was reused or re-sent for this RFQ, reset all reply state to ensure fresh sync
        if not created:
            workflow.estimator_replied = False
            workflow.estimator_reply_body = None
            workflow.estimator_replied_at = None
            workflow.detailer_replied = False
            workflow.detailer_reply_body = None
            workflow.detailer_replied_at = None
            workflow.combined_body = None
            workflow.sent_to_customer_at = None
            workflow.status = QuoteWorkflow.WorkflowStatus.FORWARDED
        workflow.save()
    except Exception as e:
        logger.error(f"Error sending email for RFQ {rfq_id}: {str(e)}", exc_info=True)

def get_next_quote_no():
    """Auto-generate next quote number, reusing gaps in sequences of the last active month."""
    qnos = list(RFQMaster.objects.filter(deleted_at__isnull=True).values_list('quote_no', flat=True))
    
    parsed = []
    for q in qnos:
        m = re.match(r'^(\d{2})-(\d{2})-(\d+)', q)
        if m:
            try:
                yy = int(m.group(1))
                mm = int(m.group(2))
                seq = int(m.group(3))
                parsed.append((yy, mm, seq))
            except ValueError:
                pass
    
    if parsed:
        parsed.sort(key=lambda x: (x[0], x[1], x[2]))
        last_yy, last_mm, last_seq = parsed[-1]
        
        # Gather all existing sequences in that specific last year and month
        existing_seqs = {seq for (yy, mm, seq) in parsed if yy == last_yy and mm == last_mm}
        
        # Find the first sequence (starting from 1) that is not currently in use
        next_seq = 1
        while next_seq in existing_seqs:
            next_seq += 1
            
        next_no = f"{last_yy:02d}-{last_mm:02d}-{next_seq:02d}"
    else:
        from django.utils import timezone
        now = timezone.now()
        next_no = f"{now.strftime('%y')}-{now.strftime('%m')}-01"
        
    return next_no


class RFQMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, CanEditRFQ]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = RFQFilter
    search_fields = ['quote_no', 'project_name', 'bid_reference', 'customer__name']
    ordering_fields = ['quote_no', 'bid_due_date', 'quote_date', 'bid_amount', 'won_lost', 'project_name', 'created_at']
    ordering = ['-quote_no']
    pagination_class = None

    def get_queryset(self):
        qs = RFQMaster.objects.select_related('customer', 'primary_estimator')
        # Exclude soft-deleted by default unless explicitly requested
        if not self.request.query_params.get('include_deleted'):
            qs = qs.filter(deleted_at__isnull=True)
        return qs

    def get_serializer_class(self):
        if self.action == 'print_view':
            return PrintSetupSerializer
        return RFQMasterSerializer

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user.username,
            updated_by=self.request.user.username,
        )

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user.username)

    def destroy(self, request, *args, **kwargs):
        """Hard delete — permanently delete from DB."""
        instance = self.get_object()
        # Clean up any QuoteWorkflow associated with this quote_no or RFQ
        QuoteWorkflow.objects.filter(
            Q(quote_id=instance.quote_no) | Q(rfq=instance)
        ).delete()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='print')
    def print_view(self, request):
        """Weekly meeting print view — filtered by bid_due_date range."""
        qs = self.get_queryset().filter(
            decision_to_bid__in=['Yes', 'Bid'],
            deleted_at__isnull=True,
        ).order_by('bid_due_date', 'customer__name')

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(bid_due_date__gte=date_from)
        if date_to:
            qs = qs.filter(bid_due_date__lte=date_to)

        serializer = PrintSetupSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='next-quote-no')
    def next_quote_no(self, request):
        return Response({'next_quote_no': get_next_quote_no()})

    @action(detail=False, methods=['post'], url_path='sync-quote-mails')
    def sync_quote_mails(self, request):
        """
        Connect to IMAP server, retrieve recent emails, check if quote-related,
        parse them, and add them to RFQMaster.
        """
        import imaplib
        import email
        from email.header import decode_header
        import sys
        import os
        sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
        import quote_helper
        from projects.models import Customer
        from django.conf import settings

        # Load configurations
        def get_setting(key, default=''):
            try:
                return SystemSetting.objects.get(key=key).value
            except SystemSetting.DoesNotExist:
                env_mapping = {
                    'email_user': ['EMAIL_HOST_USER', 'EMAIL_USER'],
                    'email_password': ['EMAIL_HOST_PASSWORD', 'EMAIL_PASSWORD'],
                    'imap_server': ['IMAP_SERVER'],
                    'imap_port': ['IMAP_PORT'],
                }
                for env_key in env_mapping.get(key, [key.upper()]):
                    val = getattr(settings, env_key, os.getenv(env_key))
                    if val:
                        if isinstance(val, str):
                            val = val.strip('\'"')
                        return val
                return default

        email_user = get_setting('email_user')
        email_password = get_setting('email_password')
        imap_server = get_setting('imap_server')

        if not imap_server:
            email_host = str(getattr(settings, 'EMAIL_HOST', os.getenv('EMAIL_HOST', ''))).lower()
            if 'zoho.in' in email_host:
                imap_server = 'imap.zoho.in'
            elif 'zoho.com' in email_host:
                imap_server = 'imap.zoho.com'
            elif 'outlook' in email_host or 'office365' in email_host:
                imap_server = 'outlook.office365.com'
            elif email_host.startswith('smtp.'):
                imap_server = email_host.replace('smtp.', 'imap.')
            else:
                imap_server = 'imap.gmail.com'

        imap_port_str = get_setting('imap_port', '993')
        try:
            imap_port = int(imap_port_str)
        except ValueError:
            imap_port = 993

        if not email_user or not email_password:
            return Response(
                {'detail': 'Email user or password not configured in system settings.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        imported_count = 0
        imported_rfqs = []
        synced_replies_count = 0

        try:
            # Connect to IMAP with fallback support for Zoho servers
            zoho_candidates = []
            if 'zoho' in str(imap_server).lower():
                zoho_servers = ['imap.zoho.com', 'imap.zoho.in', 'imappro.zoho.in', 'imappro.zoho.com']
                zoho_candidates = [imap_server] + [s for s in zoho_servers if s != imap_server]
            else:
                zoho_candidates = [imap_server]

            login_success = False
            last_err = None
            for server_candidate in zoho_candidates:
                try:
                    logger.info(f"Connecting to IMAP server: {server_candidate}:{imap_port}")
                    mail = imaplib.IMAP4_SSL(server_candidate, imap_port)
                    mail.login(email_user, email_password)
                    imap_server = server_candidate
                    login_success = True
                    break
                except imaplib.IMAP4.error as err:
                    last_err = err
                    logger.warning(f"IMAP login failed on {server_candidate}: {err}")

            if not login_success:
                raise last_err if last_err else imaplib.IMAP4.error("IMAP connection failed")

            mail.select("INBOX")

            # Search for emails
            sync_days_str = get_setting('rfq_email_sync_days', '7')
            try:
                sync_days = int(sync_days_str)
            except ValueError:
                sync_days = 7

            import datetime
            cutoff_date = (datetime.date.today() - datetime.timedelta(days=sync_days)).strftime("%d-%b-%Y")

            try:
                status_search, messages = mail.search(None, f'UNSEEN SINCE {cutoff_date}')
            except Exception:
                status_search, messages = mail.search(None, "UNSEEN")

            if status_search != "OK":
                return Response({'detail': 'Failed to search emails on IMAP server.'}, status=500)

            mail_ids = messages[0].split()
            # Fetch unread emails
            recent_mail_ids = mail_ids[-30:]

            for mail_id in recent_mail_ids:
                status_fetch, msg_data = mail.fetch(mail_id, "(RFC822)")
                if status_fetch != "OK":
                    continue

                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        
                        # Double check date in python to filter out older messages
                        date_header = msg["Date"]
                        if date_header:
                            try:
                                from email.utils import parsedate_to_datetime
                                from django.utils import timezone
                                msg_date = parsedate_to_datetime(date_header)
                                if timezone.is_aware(msg_date):
                                    cutoff_dt = timezone.now() - datetime.timedelta(days=sync_days)
                                else:
                                    cutoff_dt = datetime.datetime.now() - datetime.timedelta(days=sync_days)
                                
                                if msg_date < cutoff_dt:
                                    continue
                            except Exception as date_err:
                                logger.warning(f"Error parsing date header '{date_header}': {date_err}")
                        
                        # Extract Subject
                        subject_header = msg.get("Subject", "")
                        subject = decode_mime_header(subject_header)
                        
                        # Extract From/Sender (supports Reply-To or From)
                        from_header = msg.get("Reply-To") or msg.get("From", "")
                        sender = decode_mime_header(from_header)

                        # Extract Body
                        body = ""
                        if msg.is_multipart():
                            for part in msg.walk():
                                content_type = part.get_content_type()
                                content_disposition = str(part.get("Content-Disposition"))
                                if content_type == "text/plain" and "attachment" not in content_disposition:
                                    try:
                                        body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="replace")
                                    except Exception:
                                        pass
                                    if body and body.strip():
                                        break
                            # If no text/plain found, fallback to text/html
                            if not body or not body.strip():
                                for part in msg.walk():
                                    content_type = part.get_content_type()
                                    content_disposition = str(part.get("Content-Disposition"))
                                    if content_type == "text/html" and "attachment" not in content_disposition:
                                        try:
                                            raw_html = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", errors="replace")
                                            body = re.sub(r'<[^>]+>', ' ', raw_html)
                                            body = html.unescape(body)
                                        except Exception:
                                            pass
                                        if body and body.strip():
                                            break
                        else:
                            try:
                                raw_payload = msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8", errors="replace")
                                if msg.get_content_type() == "text/html":
                                    body = re.sub(r'<[^>]+>', ' ', raw_payload)
                                    body = html.unescape(body)
                                else:
                                    body = raw_payload
                            except Exception:
                                pass

                        subject = subject or ""
                        body = body or ""
                        if not body.strip():
                            body = subject

                        # ── 1. CHECK IF THIS EMAIL IS A REPLY TO A QUOTE WORKFLOW ──
                        ref_match = re.search(r'\[Quote Ref:\s*([^\]]+)\]', subject, re.IGNORECASE)
                        if not ref_match:
                            ref_match = re.search(r'\[Quote Ref:\s*([^\]]+)\]', body, re.IGNORECASE)

                        matched_workflow = None
                        if ref_match:
                            ref_val = ref_match.group(1).strip()
                            matched_workflow = QuoteWorkflow.objects.filter(
                                Q(quote_id__iexact=ref_val) |
                                Q(rfq__quote_no__iexact=ref_val) |
                                Q(rfq__bid_reference__iexact=ref_val)
                            ).first()

                        if not matched_workflow:
                            # Try finding quote_no (e.g. 26-08-01) in subject
                            qno_match = re.search(r'\b(\d{2}-\d{2}-\d{2,3}(?:R\d*)?)\b', subject)
                            if qno_match:
                                qno_val = qno_match.group(1).strip()
                                matched_workflow = QuoteWorkflow.objects.filter(
                                    Q(quote_id__iexact=qno_val) |
                                    Q(rfq__quote_no__iexact=qno_val)
                                ).first()

                        if matched_workflow:
                            # If linked RFQ was deleted or missing, clean up this stale workflow and skip
                            if not matched_workflow.rfq or matched_workflow.rfq.deleted_at is not None:
                                matched_workflow.delete()
                                continue

                            # Mark email as read
                            mail.store(mail_id, '+FLAGS', '\\Seen')
                            cleaned_body = quote_helper.clean_reply_body(body)
                            sender_lower = sender.lower()

                            # Determine role: detailer vs estimator / fabrication / erection
                            body_lower = body.lower()
                            det_exclusive = ['detailing cost', 'detail cost', 'detailing quote', 'detailing rate', 'drafting', 'tekla', 'model', 'ifc', 'approval drawing', 'shop drawing', 'detail']
                            est_exclusive = ['erection cost', 'erect cost', 'erection quote', 'fabrication cost', 'fabricat cost', 'fabrication quote', 'steel pricing', 'erection', 'fabrication', 'erect', 'fabricat']

                            has_det_kw = any(k in body_lower for k in det_exclusive)
                            has_est_kw = any(k in body_lower for k in est_exclusive)

                            is_detailer = False
                            if has_det_kw and not has_est_kw:
                                is_detailer = True
                            elif has_est_kw and not has_det_kw:
                                is_detailer = False
                            elif 'detailing cost' in body_lower or 'detail cost' in body_lower:
                                is_detailer = True
                            elif 'erection cost' in body_lower or 'fabrication cost' in body_lower:
                                is_detailer = False
                            else:
                                if matched_workflow.detailer_email and any(e.strip().lower() in sender_lower for e in matched_workflow.detailer_email.split(',') if e.strip()):
                                    if 'erect' in body_lower or 'fabricat' in body_lower:
                                        is_detailer = False
                                    else:
                                        is_detailer = True
                                elif matched_workflow.estimator_email and any(e.strip().lower() in sender_lower for e in matched_workflow.estimator_email.split(',') if e.strip()):
                                    if has_det_kw and not has_est_kw:
                                        is_detailer = True
                                    else:
                                        is_detailer = False
                                else:
                                    is_detailer = has_det_kw

                            from django.utils import timezone
                            now_time = timezone.now()
                            if is_detailer:
                                matched_workflow.detailer_replied = True
                                matched_workflow.detailer_reply_body = cleaned_body
                                matched_workflow.detailer_replied_at = now_time
                                if not matched_workflow.detailer_email:
                                    matched_workflow.detailer_email = sender
                            else:
                                matched_workflow.estimator_replied = True
                                matched_workflow.estimator_reply_body = cleaned_body
                                matched_workflow.estimator_replied_at = now_time
                                if not matched_workflow.estimator_email:
                                    matched_workflow.estimator_email = sender

                            # Check required scopes (Detailing, Fabrication, Erection)
                            scopes = []
                            if matched_workflow.rfq and matched_workflow.rfq.scope_of_work:
                                scopes = [s.strip().capitalize() for s in matched_workflow.rfq.scope_of_work.split(',') if s.strip()]

                            needs_det = 'Detailing' in scopes if scopes else bool(matched_workflow.detailer_email)
                            needs_est = ('Fabrication' in scopes or 'Erection' in scopes) if scopes else bool(matched_workflow.estimator_email)

                            if not needs_det and not needs_est:
                                needs_det = True
                                needs_est = True

                            is_ready_for_dispatch = False
                            if needs_det and needs_est:
                                is_ready_for_dispatch = bool(matched_workflow.estimator_replied and matched_workflow.detailer_replied)
                            elif needs_det:
                                is_ready_for_dispatch = bool(matched_workflow.detailer_replied)
                            elif needs_est:
                                is_ready_for_dispatch = bool(matched_workflow.estimator_replied)
                            else:
                                is_ready_for_dispatch = bool(matched_workflow.estimator_replied or matched_workflow.detailer_replied)

                            # Only auto-dispatch if ready and NOT already sent to customer
                            if is_ready_for_dispatch and matched_workflow.status != QuoteWorkflow.WorkflowStatus.COMPLETED and not matched_workflow.sent_to_customer_at:
                                # AUTOMATICALLY COMBINE AND SEND TO CLIENT IN BACKGROUND
                                recipient = extract_clean_email(matched_workflow.sender)
                                if not recipient and matched_workflow.rfq and matched_workflow.rfq.customer:
                                    contact = matched_workflow.rfq.customer.contacts.filter(email__isnull=False).exclude(email='').first()
                                    if contact and contact.email:
                                        recipient = extract_clean_email(contact.email)

                                if recipient:
                                    proj_name = matched_workflow.rfq.project_name if matched_workflow.rfq else ""
                                    cust_name = matched_workflow.rfq.customer.name if (matched_workflow.rfq and matched_workflow.rfq.customer) else ""
                                    combined_subject = f"Quotation Details - Ref: {matched_workflow.quote_id}" + (f" - {proj_name}" if proj_name else "")
                                    combined_body = quote_helper.build_combined_email_body(
                                        quote_id=matched_workflow.quote_id,
                                        estimator_reply=matched_workflow.estimator_reply_body,
                                        detailer_reply=matched_workflow.detailer_reply_body,
                                        project_name=proj_name,
                                        customer_name=cust_name or recipient
                                    )
                                    try:
                                        quote_helper.send_combined_quote_email(
                                            recipient_email=recipient,
                                            subject=combined_subject,
                                            body_text=combined_body
                                        )
                                        matched_workflow.status = QuoteWorkflow.WorkflowStatus.COMPLETED
                                        matched_workflow.sender = recipient
                                        matched_workflow.combined_body = combined_body
                                        matched_workflow.sent_to_customer_at = now_time
                                        logger.info(f"Auto-combined and dispatched quote for {matched_workflow.quote_id} to client {recipient}")
                                    except Exception as auto_send_err:
                                        logger.error(f"Failed to auto-send combined quote email for {matched_workflow.quote_id}: {auto_send_err}")
                                        matched_workflow.status = QuoteWorkflow.WorkflowStatus.READY
                                else:
                                    matched_workflow.status = QuoteWorkflow.WorkflowStatus.READY
                            elif not is_ready_for_dispatch and matched_workflow.status != QuoteWorkflow.WorkflowStatus.COMPLETED and not matched_workflow.sent_to_customer_at:
                                matched_workflow.status = QuoteWorkflow.WorkflowStatus.PARTIALLY_REPLIED

                            matched_workflow.save()
                            synced_replies_count += 1
                            continue

                        # ── 2. CHECK IF THIS EMAIL IS A NEW INCOMING QUOTE REQUEST ──
                        if not quote_helper.is_quote_related(subject, body):
                            continue

                        # Mark as read since it is quote related
                        mail.store(mail_id, '+FLAGS', '\\Seen')

                        # Parse quote details
                        quote_data = quote_helper.generate_quote_data(sender, subject, body)
                        quote_id = quote_data.get("quote_id")
                        bid_ref = quote_data.get("bid_reference") or quote_id
                        proj_name = quote_data.get("project_name") or "Quote Request"
                        cust_name = quote_data.get("customer_name")
                        clean_sender_email = extract_clean_email(sender)
                        msg_id = msg.get('Message-ID', '').strip()

                        if not quote_id and not bid_ref:
                            continue

                        # Check if this exact email Message-ID was already processed
                        if msg_id and QuoteWorkflow.objects.filter(message_id=msg_id).exists():
                            continue

                        # Check if RFQ already exists for the exact same project name and reference
                        existing_rfq = None
                        if bid_ref and proj_name:
                            existing_rfq = RFQMaster.objects.filter(bid_reference=bid_ref, project_name__iexact=proj_name).first()
                        if not existing_rfq and quote_id and proj_name:
                            existing_rfq = RFQMaster.objects.filter(bid_reference=quote_id, project_name__iexact=proj_name).first()

                        if existing_rfq:
                            if clean_sender_email:
                                if existing_rfq.customer:
                                    contact = existing_rfq.customer.contacts.filter(email__iexact=clean_sender_email).first()
                                    if not contact:
                                        CustomerContact.objects.create(
                                            customer=existing_rfq.customer,
                                            person=cust_name or clean_sender_email.split('@')[0],
                                            email=clean_sender_email
                                        )
                                QuoteWorkflow.objects.filter(quote_id=existing_rfq.quote_no).update(sender=clean_sender_email)
                            continue

                        # Generate next quote no
                        next_no = get_next_quote_no()

                        # Match or lookup customer
                        customer_obj = None
                        if cust_name:
                            customer_obj = Customer.objects.filter(name__iexact=cust_name.strip()).first()
                            if not customer_obj:
                                customer_obj = Customer.objects.create(
                                    name=cust_name.strip()
                                )

                        # If sender is available, ensure customer contact with this email exists
                        if customer_obj and clean_sender_email and '@' in clean_sender_email:
                            contact = customer_obj.contacts.filter(email__iexact=clean_sender_email).first()
                            if not contact:
                                CustomerContact.objects.create(
                                    customer=customer_obj,
                                    person=cust_name or clean_sender_email.split('@')[0],
                                    email=clean_sender_email
                                )

                        # Resolve budget_type
                        btype = quote_data.get("budget_type") or "Final"
                        btype = btype.strip().capitalize()
                        if btype not in ['Budget', 'Final', 'Rebid']:
                            btype = 'Final'

                        # Resolve decision_to_bid
                        dtb = quote_data.get("decision_to_bid") or "Bid"
                        dtb = dtb.strip().capitalize()
                        if dtb not in ['Yes', 'No', 'NoBid', 'Bid']:
                            dtb = 'Bid'

                        # Resolve scope_of_work
                        scope = quote_data.get("scope_of_work")
                        if scope:
                            scope_parts = [s.strip().capitalize() for s in scope.split(',') if s.strip()]
                            valid_scopes = [s for s in scope_parts if s in ['Detailing', 'Fabrication', 'Erection']]
                            if valid_scopes:
                                scope = ",".join(valid_scopes)
                            else:
                                scope = None

                        # Get or create estimator
                        estimator_obj = None
                        einit = quote_data.get("primary_estimator")
                        if einit:
                            einit = einit.strip().upper()
                            estimator_obj, _ = Estimator.objects.get_or_create(
                                initials=einit,
                                defaults={'is_active': True, 'full_name': ''}
                            )

                        # Project comments (empty if not specified in email)
                        proj_comments = quote_data.get("project_comments") or ""

                        # Create the RFQ Master record
                        rfq_rec = RFQMaster.objects.create(
                            quote_no=next_no,
                            bid_reference=bid_ref or quote_id,
                            project_name=proj_name,
                            location=quote_data.get("project_location") or "N/A",
                            bid_amount=coerce_decimal(quote_data.get("total")) or 0.0,
                            price_structure=coerce_decimal(quote_data.get("subtotal")) or 0.0,
                            project_comments=proj_comments,
                            won_lost="Pending",
                            decision_to_bid=dtb,
                            budget_type=btype,
                            customer=customer_obj,
                            bid_due_date=coerce_date(quote_data.get("bid_due_date")),
                            bid_due_time=coerce_time(quote_data.get("bid_due_time")),
                            quote_date=coerce_date(quote_data.get("date")),
                            distance_travel=coerce_decimal(quote_data.get("distance_travel")),
                            scope_of_work=scope,
                            primary_estimator=estimator_obj,
                            created_by="email_sync",
                            updated_by="email_sync"
                        )

                        # Initialize fresh QuoteWorkflow for this RFQ
                        QuoteWorkflow.objects.filter(quote_id=rfq_rec.quote_no).delete()
                        QuoteWorkflow.objects.create(
                            quote_id=rfq_rec.quote_no,
                            rfq=rfq_rec,
                            message_id=msg_id,
                            subject=subject,
                            sender=clean_sender_email or sender,
                            status=QuoteWorkflow.WorkflowStatus.FORWARDED,
                        )
                        
                        imported_count += 1
                        imported_rfqs.append({
                            'id': rfq_rec.id,
                            'quote_no': rfq_rec.quote_no,
                            'project_name': rfq_rec.project_name,
                            'bid_reference': rfq_rec.bid_reference
                        })

            mail.close()
            mail.logout()

        except imaplib.IMAP4.error as err:
            logger.warning(f"IMAP sync authentication failed for '{email_user}' on {imap_server}: {err}")
            user_hint = ""
            if "zoho" in str(imap_server).lower() or "zoho" in str(email_user).lower():
                user_hint = " Please verify IMAP access is enabled in Zoho Mail Settings (Settings -> Mail Accounts -> IMAP Access) and use a Zoho App-Specific Password if 2-Factor Authentication (2FA) is enabled."
            return Response({'detail': f'IMAP login failed for {email_user}.{user_hint}'}, status=400)
        except Exception as e:
            logger.error(f"Error during IMAP sync: {str(e)}")
            return Response({'detail': f'Error syncing email: {str(e)}'}, status=500)

        return Response({
            'success': True,
            'imported_count': imported_count,
            'imported_rfqs': imported_rfqs,
            'synced_replies_count': synced_replies_count,
            'message': f'Sync complete. {imported_count} new RFQs created, {synced_replies_count} replies updated.'
        })



    @action(detail=True, methods=['patch'], url_path='sebw-sync',
            permission_classes=[IsAuthenticated, CanEditRFQ])
    def sebw_sync(self, request, pk=None):
        """
        Accept exactly the 5 SEBW output fields.
        Audit trail: sets updated_by = 'sebw_sync' so we know which fields
        came from SEBW rather than being entered manually.
        """
        ALLOWED = {
            'bid_amount', 'quoted_profit', 'struct_erect_hours',
            'struct_erect_duration_months', 'sq_ft_structural',
        }
        data = {k: v for k, v in request.data.items() if k in ALLOWED}
        if not data:
            return Response({'detail': 'No valid SEBW fields provided.'}, status=400)

        instance = self.get_object()
        serializer = RFQMasterSerializer(
            instance, data=data, partial=True, context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(updated_by='sebw_sync')
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='duplicate',
            permission_classes=[IsAuthenticated, CanEditRFQ])
    def duplicate(self, request, pk=None):
        """
        Clone an RFQ as a Rebid with an auto-generated quote_no.
        Appends R suffix: 25-11-06 → 25-11-06R → 25-11-06R1 → 25-11-06R2 …
        Resets: won_lost=Pending, clears sfe_job_no and post-award dates.
        """
        source = self.get_object()
        # Strip any existing R suffix to get the base quote number
        base = re.sub(r'R\d*$', '', source.quote_no)
        # Try R, R1, R2 ... R19 until we find an available quote_no
        candidate = None
        for suffix in ['R'] + [f'R{i}' for i in range(1, 20)]:
            test = base + suffix
            if not RFQMaster.objects.filter(quote_no=test).exists():
                candidate = test
                break
        if not candidate:
            return Response(
                {'detail': 'Could not auto-generate rebid quote number (R through R19 all taken).'},
                status=status.HTTP_409_CONFLICT
            )

        # Clone the source record
        source.pk = None  # detach from DB to create new instance
        source.id = None
        source.quote_no = candidate
        source.budget_type = 'Rebid'
        source.won_lost = 'Pending'
        source.sfe_job_no = None
        source.awarded_job_date = None
        source.contract_executed_date = None
        source.fabrication_start_date = None
        source.awarded_amount = None
        source.deleted_at = None
        source.created_by = request.user.username
        source.updated_by = request.user.username
        source.save()
        return Response(RFQMasterSerializer(source).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='send-email',
            permission_classes=[IsAuthenticated, CanEditRFQ])
    def send_email(self, request, pk=None):
        """Send email for a single RFQ master record."""
        rfq = self.get_object()
        if not rfq.scope_of_work:
            return Response({'detail': 'Scope of Work is not selected for this project.'}, status=400)
            
        project_link = request.data.get('project_link', '').strip()
        thread = threading.Thread(target=send_rfq_email_async, args=(rfq.id, project_link))
        thread.start()
        
        return Response({'detail': 'Email sending initiated.'})

    @action(detail=False, methods=['post'], url_path='send-bulk-emails',
            permission_classes=[IsAuthenticated, CanEditRFQ])
    def send_bulk_emails(self, request):
        """Send pending emails for all non-deleted RFQs with a scope of work."""
        pending_rfqs = RFQMaster.objects.filter(
            deleted_at__isnull=True,
            scope_of_work__isnull=False,
            email_sent=False
        ).exclude(scope_of_work='')
        
        count = pending_rfqs.count()
        if count == 0:
            return Response({'detail': 'No pending emails to send.'}, status=400)
            
        for rfq in pending_rfqs:
            thread = threading.Thread(target=send_rfq_email_async, args=(rfq.id,))
            thread.start()
            
        return Response({'detail': f'Bulk sending initiated for {count} project(s).'})

    @action(detail=False, methods=['post'], url_path='test-smtp',
            permission_classes=[IsAuthenticated, CanEditRFQ])
    def test_smtp(self, request):
        """Diagnose SMTP configurations by sending a test email synchronously."""
        try:
            recipient = request.data.get('email', 'support@caldimengg.in')
            subject = "SFE SMTP Diagnostics Test"
            body = "This is a synchronous test email to verify SMTP configuration on the SFE Milestone server."
            
            send_mail(
                subject=subject,
                message=body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[recipient],
                fail_silently=False
            )
            return Response({
                'success': True,
                'detail': f'Test email sent successfully to {recipient}.',
                'smtp_host': settings.EMAIL_HOST,
                'smtp_port': settings.EMAIL_PORT,
                'from_email': settings.DEFAULT_FROM_EMAIL
            })
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            return Response({
                'success': False,
                'detail': str(e),
                'traceback': error_trace,
                'smtp_host': getattr(settings, 'EMAIL_HOST', None),
                'smtp_port': getattr(settings, 'EMAIL_PORT', None),
                'from_email': getattr(settings, 'DEFAULT_FROM_EMAIL', None)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='upload-excel',
            permission_classes=[IsAuthenticated, CanEditRFQ],
            parser_classes=[MultiPartParser, FormParser])
    def upload_excel(self, request):
        """
        Accept an .xlsx file upload.
        ?commit=true  → actually save records
        ?commit=false → dry-run preview only (default)
        Returns { new: [], updated: [], skipped: int, errors: [] }
        """
        import sys, os
        from datetime import datetime, date, time
        import datetime as dt
        from decimal import Decimal, InvalidOperation
        from django.db import transaction

        file_obj = request.FILES.get('file')
        commit = request.data.get('commit', 'false').lower() == 'true'

        if not file_obj:
            return Response({'detail': 'No file uploaded.'}, status=400)
        if not file_obj.name.endswith('.xlsx'):
            return Response({'detail': 'Only .xlsx files are accepted.'}, status=400)

        try:
            import openpyxl
        except ImportError:
            return Response({'detail': 'openpyxl not installed on server.'}, status=500)

        # Full column mapping from migrate_excel.py
        EXCEL_COLUMN_MAP = {
            'A': 'quote_no',
            'E': 'budget_type',
            'F': 'bid_reference',
            'G': 'project_name',
            'H': 'project_comments',
            'I': 'bid_due_date',
            'J': 'bid_due_time',
            'K': 'location',
            'L': 'distance_travel',
            'M': 'aisc_fab',
            'N': 'aisc_erect',
            'O': 'domestic_steel',
            'P': 'leed_project',
            'Q': 'minority_participation',
            'R': 'prevailing_wage',
            'S': 'ccip_ocip',
            'T': 'bonded',
            'U': 'paint',
            'V': 'galvanised',
            'W': 'professional_engineer',
            'X': 'third_party_inspection',
            'Y': 'tax_status',
            'Z': '_customer_name',
            'AA': 'decision_to_bid',
            'AB': '_estimator_initials',
            'AC': 'outsourced_estimator',
            'AD': 'sent_to_jd',
            'AE': 'sent_to_detailing',
            'AF': 'sent_to_erection',
            'AG': 'est_sqft_ton',
            'AH': 'price_structure',
            'AI': 'price_erection',
            'AJ': 'price_misc',
            'AK': 'price_misc_erection',
            'AL': 'bid_amount',
            'AM': 'quoted_profit',
            'AN': 'ton_steel',
            'AO': 'ton_joist',
            'AQ': 'num_main_structural_pcs',
            'AS': 'sq_ft_structural',
            'AZ': 'struct_fab_hours',
            'BA': 'struct_fab_start_month',
            'BB': 'struct_fab_duration_months',
            'BE': 'misc_fab_hours',
            'BF': 'misc_fab_start_month',
            'BG': 'misc_fab_duration_months',
            'BJ': 'struct_erect_hours',
            'BK': 'struct_erect_start_month',
            'BL': 'struct_erect_duration_months',
            'BO': 'misc_erect_hours',
            'BP': 'misc_erect_start_month',
            'BQ': 'misc_erect_duration_months',
            'BT': 'estimating_hours',
            'BU': 'quote_date',
            'BV': 'won_lost',
            'BW': 'follow_up_notes',
            'BX': 'follow_up_date',
            'BY': 'awarded_amount',
            'BZ': 'sfe_job_no',
            'CA': 'awarded_job_date',
            'CB': 'contract_executed_date',
            'CC': 'fabrication_start_date',
        }

        BOOL_FIELDS = {
            'aisc_fab', 'aisc_erect', 'domestic_steel', 'leed_project',
            'minority_participation', 'prevailing_wage', 'ccip_ocip', 'bonded',
            'paint', 'galvanised', 'professional_engineer', 'third_party_inspection',
            'sent_to_jd', 'sent_to_detailing', 'sent_to_erection',
        }

        DATE_FIELDS = {
            'bid_due_date', 'quote_date', 'follow_up_date',
            'awarded_job_date', 'contract_executed_date', 'fabrication_start_date',
            'struct_fab_start_month', 'misc_fab_start_month',
            'struct_erect_start_month', 'misc_erect_start_month',
        }

        DECIMAL_FIELDS = {
            'price_structure', 'price_erection', 'price_misc', 'price_misc_erection',
            'bid_amount', 'quoted_profit', 'awarded_amount',
            'ton_steel', 'ton_joist', 'sq_ft_structural',
            'struct_fab_hours', 'misc_fab_hours', 'struct_erect_hours', 'misc_erect_hours',
            'estimating_hours', 'distance_travel',
        }

        INT_FIELDS = {
            'num_main_structural_pcs', 'sfe_job_no',
            'struct_fab_duration_months', 'misc_fab_duration_months',
            'struct_erect_duration_months', 'misc_erect_duration_months',
        }

        WON_LOST_MAP = {
            'Won': 'Won', 'won': 'Won', 'W': 'Won',
            'Lost': 'Lost', 'lost': 'Lost', 'L': 'Lost',
            'Pending': 'Pending', 'pending': 'Pending', 'P': 'Pending',
            '': 'Pending', None: 'Pending',
        }

        def col_idx(letter):
            result = 0
            for c in letter.upper():
                result = result * 26 + (ord(c) - ord('A') + 1)
            return result - 1



        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        if 'Data' not in wb.sheetnames:
            return Response({'detail': "Sheet 'Data' not found in workbook."}, status=400)
        ws = wb['Data']

        col_positions = {letter: col_idx(letter) for letter in EXCEL_COLUMN_MAP}

        new_records, updated_records, skipped, errors = [], [], 0, []
        parsed_records = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            raw_qno = row[col_positions['A']] if col_positions['A'] < len(row) else None
            if not raw_qno:
                skipped += 1
                continue
            quote_no = str(raw_qno).strip()
            
            # Normalize quote number pattern YY-MM-SEQ
            if not re.match(r'^\d{2}-\d{2}-\d{2,3}(R\d?)?$', quote_no):
                quote_no = re.sub(r'\s+', '', quote_no)
                if not re.match(r'^\d{2}-\d{2}-\d{2,3}(R\d?)?$', quote_no):
                    skipped += 1
                    continue

            # Build record dict
            rec: dict = {'quote_no': quote_no}
            
            for letter, field in EXCEL_COLUMN_MAP.items():
                if field.startswith('_'):
                    continue  # customer / estimator handled separately
                
                idx = col_positions[letter]
                raw = row[idx] if idx < len(row) else None
                
                if field in BOOL_FIELDS:
                    rec[field] = coerce_bool(raw)
                elif field in DATE_FIELDS:
                    rec[field] = coerce_date(raw)
                elif field in DECIMAL_FIELDS:
                    rec[field] = coerce_decimal(raw)
                elif field in INT_FIELDS:
                    rec[field] = coerce_int(raw)
                elif field == 'won_lost':
                    rec[field] = WON_LOST_MAP.get(raw, 'Pending')  # type: ignore
                elif field == 'bid_due_time':
                    rec[field] = coerce_time(raw)
                else:
                    rec[field] = str(raw).strip() if raw is not None else ''

            # Extract customer name
            cname = str(row[col_positions['Z']] if col_positions['Z'] < len(row) else '').strip()
            if cname:
                rec['_customer_name'] = cname

            # Extract estimator initials
            einit = str(row[col_positions['AB']] if col_positions['AB'] < len(row) else '').strip()
            if einit:
                rec['_estimator_initials'] = einit

            parsed_records.append(rec)

        # Categorize records for response preview
        for rec in parsed_records:
            exists = RFQMaster.objects.filter(quote_no=rec['quote_no']).exists()
            entry = {
                'quote_no': rec['quote_no'],
                'project_name': rec.get('project_name', ''),
                'is_new': not exists,
            }
            if exists:
                updated_records.append(entry)
            else:
                new_records.append(entry)

        if commit:
            seen_job_nos = set()
            try:
                with transaction.atomic():
                    for rec in parsed_records:
                        quote_no = rec['quote_no']
                        
                        # Get or create customer
                        # Attempt to find an existing customer by name; do NOT create new customers from Excel data
                        customer_obj = None
                        cname = rec.get('_customer_name')
                        if cname:
                            try:
                                customer_obj = Customer.objects.get(name=cname)
                            except Customer.DoesNotExist:
                                # If customer not found in master, leave as None (will be ignored or cause validation error)
                                customer_obj = None
                        # Get or create estimator
                        estimator_obj = None
                        einit = rec.get('_estimator_initials')
                        if einit:
                            estimator_obj, _ = Estimator.objects.get_or_create(
                                initials=einit,
                                defaults={'is_active': True, 'full_name': ''}
                            )
                            
                        # Build RFQ field dictionary for update/create
                        defaults = {k: v for k, v in rec.items() if not k.startswith('_')}
                        defaults['customer'] = customer_obj
                        defaults['primary_estimator'] = estimator_obj
                        
                        # Handle duplicate sfe_job_no safely to avoid database IntegrityError
                        jno = defaults.get('sfe_job_no')
                        if jno is not None:
                            if jno in seen_job_nos or RFQMaster.objects.filter(sfe_job_no=jno).exclude(quote_no=quote_no).exists():
                                defaults['sfe_job_no'] = None
                            else:
                                seen_job_nos.add(jno)

                        # Save or update RFQMaster instance
                        rfq_inst = RFQMaster.objects.filter(quote_no=quote_no).first()
                        if rfq_inst:
                            # Update existing
                            for k, v in defaults.items():
                                setattr(rfq_inst, k, v)
                            rfq_inst.updated_by = str(request.user.username or 'excel_upload')
                            rfq_inst.save()
                        else:
                            # Create new
                            rfq_inst = RFQMaster(**defaults)
                            rfq_inst.created_by = str(request.user.username or 'excel_upload')  # type: ignore
                            rfq_inst.updated_by = str(request.user.username or 'excel_upload')  # type: ignore
                            rfq_inst.save()
            except Exception as e:
                return Response({
                    'detail': f'Error occurred during import transaction: {str(e)}'
                }, status=500)

        result = {
            'new': new_records,
            'updated': updated_records,
            'skipped': skipped,
            'errors': errors,
            'commit': commit,
        }

        if commit:
            result['message'] = f'Successfully committed {len(new_records)} new and {len(updated_records)} updated records to the database.'

        return Response(result)



# ─────────────────────────────────────────────────────────────────────────────
# LOOKUP TABLES
# ─────────────────────────────────────────────────────────────────────────────

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all().order_by('name')
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']
    pagination_class = None  # Return all customers for dropdowns


class EstimatorViewSet(viewsets.ModelViewSet):
    queryset = Estimator.objects.filter(is_active=True).order_by('initials')
    serializer_class = EstimatorSerializer
    permission_classes = [IsAuthenticated, IsManagerOrReadOnly]
    pagination_class = None


class MonthlyBidGoalViewSet(viewsets.ModelViewSet):
    queryset = MonthlyBidGoal.objects.all().order_by('year', 'month')
    serializer_class = MonthlyBidGoalSerializer
    permission_classes = [IsAuthenticated, CanEditRFQ]
    pagination_class = None

    @action(detail=False, methods=['get'])
    def current_year(self, request):
        """Return goals for current year, auto-creating missing months at $2M."""
        from django.utils import timezone
        year = timezone.now().year
        goals = []
        for month in range(1, 13):
            obj, _ = MonthlyBidGoal.objects.get_or_create(
                year=year, month=month,
                defaults={'goal': 2000000}
            )
            goals.append(obj)
        return Response(MonthlyBidGoalSerializer(goals, many=True).data)


class SystemSettingViewSet(viewsets.ModelViewSet):
    queryset = SystemSetting.objects.all()
    serializer_class = SystemSettingSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None


class QuoteWorkflowViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, CanEditRFQ]
    serializer_class = QuoteWorkflowSerializer
    queryset = QuoteWorkflow.objects.select_related('rfq', 'rfq__customer').all().order_by('-created_at')
    pagination_class = None

    def perform_destroy(self, instance):
        rfq = instance.rfq
        instance.delete()
        if rfq and rfq.created_by == 'email_sync':
            rfq.delete()

    @action(detail=True, methods=['get'], url_path='preview-combined')
    def preview_combined(self, request, pk=None):
        """Build and return the preview for combined estimation & detailing email."""
        workflow = self.get_object()
        project_name = workflow.rfq.project_name if workflow.rfq else ""
        customer_name = workflow.rfq.customer.name if (workflow.rfq and workflow.rfq.customer) else ""
        
        subject = f"Quotation Details - Ref: {workflow.quote_id}" + (f" - {project_name}" if project_name else "")
        body = quote_helper.build_combined_email_body(
            quote_id=workflow.quote_id,
            estimator_reply=workflow.estimator_reply_body,
            detailer_reply=workflow.detailer_reply_body,
            project_name=project_name,
            customer_name=customer_name or workflow.sender
        )
        
        recipient = extract_clean_email(workflow.sender)
        if not recipient and workflow.rfq and workflow.rfq.customer:
            contact = workflow.rfq.customer.contacts.filter(email__isnull=False).exclude(email='').first()
            if contact and contact.email:
                recipient = extract_clean_email(contact.email)

        # Dynamic SFE dispatcher email address
        from_email = getattr(settings, 'EMAIL_HOST_USER', os.getenv('EMAIL_HOST_USER', 'thamizh1700@gmail.com'))
        try:
            from_email = SystemSetting.objects.get(key='email_user').value
        except Exception:
            pass

        return Response({
            'quote_id': workflow.quote_id,
            'recipient': recipient or '',
            'from_email': from_email or 'thamizh1700@gmail.com',
            'subject': subject,
            'body': body,
            'estimator_replied': workflow.estimator_replied,
            'estimator_reply_body': workflow.estimator_reply_body,
            'detailer_replied': workflow.detailer_replied,
            'detailer_reply_body': workflow.detailer_reply_body,
            'status': workflow.status,
            'is_ready': bool(workflow.estimator_replied and workflow.detailer_replied),
        })

    @action(detail=True, methods=['post'], url_path='combine-and-send')
    def combine_and_send(self, request, pk=None):
        """Combine replies and dispatch quotation email to customer."""
        workflow = self.get_object()
        project_name = workflow.rfq.project_name if workflow.rfq else ""
        customer_name = workflow.rfq.customer.name if (workflow.rfq and workflow.rfq.customer) else ""
        
        recipient = extract_clean_email(request.data.get('recipient')) or extract_clean_email(workflow.sender)
        if not recipient and workflow.rfq and workflow.rfq.customer:
            contact = workflow.rfq.customer.contacts.filter(email__isnull=False).exclude(email='').first()
            if contact and contact.email:
                recipient = extract_clean_email(contact.email)

        if not recipient:
            return Response({'detail': 'No recipient customer email address specified.'}, status=400)

        subject = request.data.get('subject') or (f"Quotation Details - Ref: {workflow.quote_id}" + (f" - {project_name}" if project_name else ""))
        body = request.data.get('body') or quote_helper.build_combined_email_body(
            quote_id=workflow.quote_id,
            estimator_reply=workflow.estimator_reply_body,
            detailer_reply=workflow.detailer_reply_body,
            project_name=project_name,
            customer_name=customer_name or recipient
        )

        try:
            quote_helper.send_combined_quote_email(
                recipient_email=recipient,
                subject=subject,
                body_text=body
            )
            from django.utils import timezone
            workflow.status = QuoteWorkflow.WorkflowStatus.COMPLETED
            workflow.sender = recipient
            workflow.combined_body = body
            workflow.sent_to_customer_at = timezone.now()
            workflow.save()

            return Response({
                'success': True,
                'message': f'Combined quotation email sent successfully to {recipient}.',
                'workflow': QuoteWorkflowSerializer(workflow).data
            })
        except Exception as e:
            logger.error(f"Failed to send combined quote email: {str(e)}", exc_info=True)
            return Response({'detail': f'Failed to send email: {str(e)}'}, status=500)

    @action(detail=False, methods=['post'], url_path='sync-replies')
    def sync_replies(self, request):
        """Trigger sync of reply emails from IMAP account."""
        rfq_viewset = RFQMasterViewSet()
        rfq_viewset.request = request
        rfq_viewset.format_kwarg = None
        return rfq_viewset.sync_quote_mails(request)


# ─────────────────────────────────────────────────────────────────────────────
# BACKGROUND DAEMON RUNNER FOR AUTOMATIC QUOTE REPLY COMBINATION & DISPATCH
# ─────────────────────────────────────────────────────────────────────────────

_bg_daemon_started = False

def start_quote_sync_background_daemon():
    """Starts a background daemon thread that periodically checks for quote replies and auto-dispatches."""
    global _bg_daemon_started
    if _bg_daemon_started:
        return
    _bg_daemon_started = True

    import time
    def _worker():
        time.sleep(10)  # Initial delay after server boot
        consecutive_auth_failures = 0
        while True:
            # Respect env toggle to disable background polling if requested
            if str(os.getenv('ENABLE_BACKGROUND_IMAP_SYNC', 'true')).lower() in ('false', '0', 'no', 'off'):
                time.sleep(300)
                continue

            try:
                viewset = RFQMasterViewSet()
                class MockRequest:
                    GET = {}
                    data = {}
                viewset.request = MockRequest()
                viewset.format_kwarg = None
                res = viewset.sync_quote_mails(viewset.request)
                if hasattr(res, 'data') and isinstance(res.data, dict) and 'IMAP login failed' in str(res.data.get('detail', '')):
                    consecutive_auth_failures += 1
                else:
                    consecutive_auth_failures = 0
            except Exception as e:
                logger.debug(f"Background quote sync cycle error: {e}")

            # Back off polling interval if authentication fails repeatedly
            if consecutive_auth_failures >= 2:
                time.sleep(600)  # Back off 10 mins on persistent auth failure
            else:
                time.sleep(45)  # Normal polling interval

    t = threading.Thread(target=_worker, daemon=True, name="QuoteWorkflowBackgroundSync")
    t.start()



